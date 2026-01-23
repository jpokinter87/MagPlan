"""
Génération des lignes de permanences
"""

from datetime import datetime
import openpyxl

from .config import (
    PERMANENCES, COULEUR_LEGENDE_PERMANENCES, COULEUR_LEGENDE_PERM_NUIT,
    COULEUR_LEGENDE_DEBATS_JLD, BLANC, GRIS, ROUGE, BLEU_FONCE
)
from .dates import est_jour_non_ouvre
from .styles import appliquer_bordures, coloriser_legende, appliquer_style_cellule


def ajouter_permanences(ws, ligne_actuelle, mois, annee, nb_jours):
    """
    Ajoute les lignes de permanences (lignes 2-11)

    Returns:
        int: La ligne actuelle après ajout des permanences
    """
    for idx_perm, perm in enumerate(PERMANENCES):
        ws.merge_cells(start_row=ligne_actuelle, start_column=1,
                      end_row=ligne_actuelle, end_column=2)
        ws.cell(ligne_actuelle, 1, perm['libelle'])

        # Coloriser la légende
        coloriser_legende(ws, ligne_actuelle, COULEUR_LEGENDE_PERMANENCES, nb_jours)

        # Coloriser les cellules de dates
        for jour in range(1, nb_jours + 1):
            col = 2 + jour
            col_letter = openpyxl.utils.get_column_letter(col)
            date = datetime(annee, mois, jour)

            cell = ws.cell(ligne_actuelle, col)

            # PERMANENCE HIÉRARCHIQUE (index 0)
            if idx_perm == 0:
                if date.weekday() == 6:  # Dimanche
                    col_letter_samedi = openpyxl.utils.get_column_letter(col - 1)
                    cell.value = f'={col_letter_samedi}{ligne_actuelle}'
                    appliquer_style_cellule(cell, ROUGE)
                elif est_jour_non_ouvre(date):
                    appliquer_style_cellule(cell, ROUGE)
                else:
                    appliquer_style_cellule(cell, perm['couleur'])

            # PERMANENCES PAP A et B (indices 1 et 2)
            elif idx_perm in [1, 2]:
                if est_jour_non_ouvre(date):
                    appliquer_style_cellule(cell, BLANC)
                else:
                    appliquer_style_cellule(cell, perm['couleur'])

            # Autres permanences (indices 3-9)
            else:
                if est_jour_non_ouvre(date):
                    appliquer_style_cellule(cell, GRIS)
                else:
                    appliquer_style_cellule(cell, perm['couleur'])

            appliquer_bordures(cell)

        ligne_actuelle += 1

    return ligne_actuelle


def ajouter_debats_jld(ws, ligne_actuelle, nb_jours):
    """
    Ajoute la ligne "Débats JLD" qui copie automatiquement la permanence de nuit
    Légende ET cellules en bleu foncé

    Returns:
        int: La ligne actuelle après ajout
    """
    ws.merge_cells(start_row=ligne_actuelle, start_column=1,
                  end_row=ligne_actuelle, end_column=2)
    ws.cell(ligne_actuelle, 1, 'Débats JLD')

    # Coloriser la légende en bleu foncé
    coloriser_legende(ws, ligne_actuelle, COULEUR_LEGENDE_DEBATS_JLD, nb_jours)

    # Bordure épaisse en haut
    for col in range(1, 3 + nb_jours):
        cell = ws.cell(ligne_actuelle, col)
        appliquer_bordures(cell, top='medium')

    ligne_perm_nuit = ligne_actuelle + 1

    for jour in range(1, nb_jours + 1):
        col = 2 + jour
        col_letter = openpyxl.utils.get_column_letter(col)
        cell = ws.cell(ligne_actuelle, col)

        # Formule : copier la permanence de nuit DU MÊME JOUR
        cell.value = f'={col_letter}{ligne_perm_nuit}'
        # Fond bleu foncé avec police blanche
        appliquer_style_cellule(cell, BLEU_FONCE)
        appliquer_bordures(cell, top='medium')

    ligne_actuelle += 1
    return ligne_actuelle


def ajouter_permanence_nuit(ws, ligne_actuelle, mois, annee, nb_jours):
    """
    Ajoute la ligne "Permanence de Nuit"

    Returns:
        int: La ligne actuelle après ajout
    """
    ws.merge_cells(start_row=ligne_actuelle, start_column=1,
                  end_row=ligne_actuelle, end_column=2)
    ws.cell(ligne_actuelle, 1, 'Permanence de Nuit')
    coloriser_legende(ws, ligne_actuelle, COULEUR_LEGENDE_PERM_NUIT, nb_jours)

    for jour in range(1, nb_jours + 1):
        col = 2 + jour
        date = datetime(annee, mois, jour)

        cell = ws.cell(ligne_actuelle, col)
        # Bleu soutenu avec police blanche
        appliquer_style_cellule(cell, 'FF4472C4')
        appliquer_bordures(cell)

    # Bordure épaisse en bas
    for col in range(1, 3 + nb_jours):
        cell = ws.cell(ligne_actuelle, col)
        appliquer_bordures(cell, bottom='medium')

    ligne_actuelle += 1
    return ligne_actuelle