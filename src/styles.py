"""
Styles, couleurs et formatage Excel
"""

from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from typing import Optional, Literal

BorderStyle = Literal["thin", "medium", "thick", "double", "dotted", "dashed"]

# Police par défaut pour tout le document
POLICE_DEFAUT = "Calibri"  # ou "Arial"
TAILLE_POLICE_DEFAUT = 11


def est_couleur_sombre(couleur_hex):
    """
    Détermine si une couleur est sombre (nécessite police blanche)
    Utilise la formule de luminance relative

    Args:
        couleur_hex: Couleur au format 'FFRRGGBB'

    Returns:
        bool: True si la couleur est sombre
    """
    # Extraire les composantes RGB (ignorer les 2 premiers caractères FF)
    if len(couleur_hex) == 8:
        r = int(couleur_hex[2:4], 16)
        g = int(couleur_hex[4:6], 16)
        b = int(couleur_hex[6:8], 16)
    else:
        return False

    # Calculer la luminance relative (formule W3C)
    luminance = (0.299 * r + 0.587 * g + 0.114 * b) / 255

    # Si luminance < 0.5, la couleur est sombre
    return luminance < 0.5


def appliquer_style_cellule(cell, couleur_fond, texte=None, gras=False):
    """
    Applique un style complet à une cellule : fond, police, alignement

    Args:
        cell: Cellule Excel
        couleur_fond: Couleur de fond au format 'FFRRGGBB'
        texte: Texte optionnel à mettre dans la cellule
        gras: Police en gras
    """
    # Appliquer le fond
    cell.fill = PatternFill(start_color=couleur_fond, end_color=couleur_fond, fill_type='solid')

    # Déterminer la couleur de la police (blanc si fond sombre, noir sinon)
    if est_couleur_sombre(couleur_fond):
        couleur_police = 'FFFFFFFF'  # Blanc
    else:
        couleur_police = 'FF000000'  # Noir

    # Appliquer la police
    cell.font = Font(
        name=POLICE_DEFAUT,
        color=couleur_police,
        bold=gras,
        size=11
    )

    # Appliquer l'alignement centré avec réduction automatique
    cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=False,
        shrink_to_fit=True  # Réduire la taille si texte trop long
    )

    # Définir le texte si fourni
    if texte is not None:
        cell.value = str(texte).upper()  # Forcer les majuscules


def appliquer_alignement_centre(cell):
    """
    Applique uniquement l'alignement centré à une cellule
    (sans modifier le style existant)
    """
    cell.alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=False,
        shrink_to_fit=True
    )


def appliquer_police_standard(cell, gras=False, taille=None):
    """
    Applique la police standard sans modifier la couleur existante

    Args:
        cell: Cellule Excel
        gras: Police en gras
        taille: Taille de la police (par défaut TAILLE_POLICE_DEFAUT)
    """
    # Garder la couleur actuelle si elle existe
    couleur_actuelle = cell.font.color if cell.font and cell.font.color else None

    cell.font = Font(
        name=POLICE_DEFAUT,
        color=couleur_actuelle,
        bold=gras,
        size=taille if taille else TAILLE_POLICE_DEFAUT
    )


def appliquer_bordures(cell,
                      top: Optional[BorderStyle] = 'thin',
                      bottom: Optional[BorderStyle] = 'thin',
                      left: Optional[BorderStyle] = 'thin',
                      right: Optional[BorderStyle] = 'thin'):
    """Applique des bordures à une cellule"""
    border = Border(
        top=Side(style=top) if top else None,
        bottom=Side(style=bottom) if bottom else None,
        left=Side(style=left) if left else None,
        right=Side(style=right) if right else None
    )
    cell.border = border


def coloriser_legende(ws, ligne, couleur, nb_jours):
    """Colorise les colonnes A-B pour la légende et applique des bordures"""
    from .config import BLANC

    # Colonne A avec couleur et police adaptée
    cell_a = ws.cell(ligne, 1)
    appliquer_style_cellule(cell_a, couleur, gras=False)
    appliquer_bordures(cell_a)

    # Colonne B blanche
    cell_b = ws.cell(ligne, 2)
    appliquer_style_cellule(cell_b, BLANC, gras=False)
    appliquer_bordures(cell_b)

    # Bordures sur toutes les cellules de dates
    for jour in range(1, nb_jours + 1):
        cell = ws.cell(ligne, 2 + jour)
        appliquer_bordures(cell)