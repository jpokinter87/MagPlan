"""
Orchestration de la génération du planning
"""

import os
from calendar import monthrange
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

from .audiences import ajouter_toutes_audiences, ajouter_assises_et_ccd
from .config import (
    DOSSIER_DATA, DOSSIER_SORTIES, FICHIER_REPARTITION_DEFAUT,
    MOIS_FR, JAUNE_CLAIR
)
from .dates import est_jour_ferie
from .permanences import ajouter_permanences, ajouter_debats_jld, ajouter_permanence_nuit
from .repartition import charger_tableau_repartition
from .styles import (
    appliquer_bordures, coloriser_legende, appliquer_alignement_centre, appliquer_police_standard,
    POLICE_DEFAUT, TAILLE_POLICE_DEFAUT
)


def creer_dossiers():
    """Crée les dossiers data/ et sorties/ s'ils n'existent pas"""
    if not os.path.exists(DOSSIER_DATA):
        os.makedirs(DOSSIER_DATA)
        print(f"ℹ️  Dossier '{DOSSIER_DATA}/' créé")

    if not os.path.exists(DOSSIER_SORTIES):
        os.makedirs(DOSSIER_SORTIES)
        print(f"ℹ️  Dossier '{DOSSIER_SORTIES}/' créé")


def trouver_fichier_repartition():
    """Trouve le fichier de répartition dans différents emplacements possibles"""
    chemins_possibles = [
        os.path.join(DOSSIER_DATA, FICHIER_REPARTITION_DEFAUT),
        os.path.join(DOSSIER_DATA, "tableau_répartition_audiences.xlsx"),
        os.path.join(DOSSIER_DATA, "tableau repartition audiences.xlsx"),
        FICHIER_REPARTITION_DEFAUT,
        "tableau_répartition_audiences.xlsx",
        "tableau repartition audiences.xlsx"
    ]

    for chemin in chemins_possibles:
        if os.path.exists(chemin):
            return chemin

    return None


def generer_nom_fichier(mois, annee):
    """Génère un nom de fichier au format: MM - MOIS ANNEE.xlsx"""
    mois_texte = MOIS_FR[mois - 1]
    nom_fichier = f"{mois:02d} - {mois_texte} {annee}.xlsx"
    return nom_fichier


def ajouter_ligne_dates(ws, mois, annee, nb_jours):
    """Ajoute la ligne 1 avec les dates, centrées et en sans-serif"""
    for jour in range(1, nb_jours + 1):
        date = datetime(annee, mois, jour)
        col = 2 + jour
        cell = ws.cell(1, col, date)
        cell.number_format = 'ddd d mmm'

        # Appliquer la police standard
        cell.font = Font(
            name=POLICE_DEFAUT,
            size=TAILLE_POLICE_DEFAUT,
            bold=False
        )

        # Centrer le texte
        appliquer_alignement_centre(cell)

        # Bordures
        appliquer_bordures(cell, top='medium')

        # Colorer les jours fériés en rouge clair
        if est_jour_ferie(date):
            cell.fill = PatternFill(start_color='FFFFCCCC', end_color='FFFFCCCC', fill_type='solid')

    # Cellules A1-B1
    for col in [1, 2]:
        cell = ws.cell(1, col)
        appliquer_police_standard(cell)
        appliquer_alignement_centre(cell)
        appliquer_bordures(cell, top='medium')


def ajouter_ligne_indisponibles(ws, ligne_actuelle, nb_jours):
    """
    Ajoute la ligne "Indisponibles" avec formules automatiques

    Returns:
        int: La ligne actuelle après ajout
    """
    from .styles import appliquer_style_cellule
    from .config import JAUNE_CLAIR

    ws.cell(ligne_actuelle, 1, 'Indisponibles')
    ws.cell(ligne_actuelle, 2, '(auto)')
    coloriser_legende(ws, ligne_actuelle, JAUNE_CLAIR, nb_jours)

    ligne_perm_nuit = 13
    ligne_perm_hier = 2

    for jour in range(1, nb_jours + 1):
        col = 2 + jour
        col_letter = openpyxl.utils.get_column_letter(col)
        date_col = f"{col_letter}1"

        col_letter_samedi = openpyxl.utils.get_column_letter(col - 2) if col > 3 else ""
        col_letter_dimanche = openpyxl.utils.get_column_letter(col - 1) if col > 2 else ""

        # Formule
        formule = (
            f'=IF(WEEKDAY({date_col},2)=1,'
            f'TEXTJOIN(", ",TRUE,'
            f'{col_letter_samedi}{ligne_perm_hier},'
            f'{col_letter_dimanche}{ligne_perm_hier},'
            f'{col_letter_dimanche}{ligne_perm_nuit}'
            f'),'
            f'IF(COLUMN({col_letter}1)>2,{col_letter_dimanche}{ligne_perm_nuit},"")'
            f')'
        )

        cell = ws.cell(ligne_actuelle, col)
        cell.value = formule
        appliquer_style_cellule(cell, JAUNE_CLAIR)
        appliquer_bordures(cell)

    ligne_actuelle += 1
    return ligne_actuelle, ligne_actuelle - 1


def ajuster_mise_en_page(ws, nb_jours):
    """Ajuste les largeurs de colonnes et fige les volets"""
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25

    for jour in range(1, nb_jours + 1):
        col_letter = openpyxl.utils.get_column_letter(2 + jour)
        ws.column_dimensions[col_letter].width = 18

    # Figer les 2 premières colonnes
    ws.freeze_panes = 'C1'


def creer_planning_mensuel(mois, annee, fichier_repartition, fichier_sortie):
    """Crée le planning mensuel complet"""

    print(f"📅 Génération du planning pour {MOIS_FR[mois-1]} {annee}...")

    # Charger les règles de répartition
    repartition = charger_tableau_repartition(fichier_repartition)

    # Créer un nouveau workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Feuil1'

    # Configuration générale
    nb_jours = monthrange(annee, mois)[1]
    ligne_actuelle = 1

    # Ligne 1 : Dates
    ajouter_ligne_dates(ws, mois, annee, nb_jours)
    ligne_actuelle = 2

    # Permanences (lignes 2-11)
    ligne_actuelle = ajouter_permanences(ws, ligne_actuelle, mois, annee, nb_jours)

    # Débats JLD (ligne 12)
    ligne_actuelle = ajouter_debats_jld(ws, ligne_actuelle, nb_jours)

    # Permanence de nuit (ligne 13)
    ligne_actuelle = ajouter_permanence_nuit(ws, ligne_actuelle, mois, annee, nb_jours)

    # Audiences (toutes les audiences)
    ligne_actuelle = ajouter_toutes_audiences(ws, ligne_actuelle, mois, annee, nb_jours, repartition)

    # Assises et CCD
    ligne_actuelle = ajouter_assises_et_ccd(ws, ligne_actuelle, nb_jours)

    # Ligne Indisponibles
    ligne_actuelle, ligne_indisponibles = ajouter_ligne_indisponibles(ws, ligne_actuelle, nb_jours)

    # Ajuster la mise en page
    ajuster_mise_en_page(ws, nb_jours)

    # Sauvegarder
    wb.save(fichier_sortie)
    print(f"✅ Planning généré: {fichier_sortie}")