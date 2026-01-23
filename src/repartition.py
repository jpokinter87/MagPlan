"""
Chargement et traitement du tableau de répartition des audiences
"""

import openpyxl
from .config import CODES_COULEURS


def charger_tableau_repartition(fichier_repartition):
    """Charge le tableau de répartition et extrait les règles d'attribution"""
    wb = openpyxl.load_workbook(fichier_repartition)
    ws = wb.active

    repartition_audiences = {}

    # Parcourir les lignes d'audiences (lignes 4 à 33)
    for row in range(4, 34):
        horaire = ws.cell(row, 1).value
        type_audience = ws.cell(row, 2).value

        if not type_audience:
            continue

        cle = f"{horaire}_{type_audience}"
        repartition_audiences[cle] = {
            'LUNDI': {}, 'MARDI': {}, 'MERCREDI': {}, 'JEUDI': {}, 'VENDREDI': {}
        }

        # Colonnes par jour
        jour_cols = {
            'LUNDI': [3, 4, 5, 6, 7],
            'MARDI': [8, 9, 10, 11, 12],
            'MERCREDI': [13, 14, 15, 16, 17],
            'JEUDI': [18, 19, 20, 21, 22],
            'VENDREDI': [23, 24, 25, 26, 27]
        }

        for jour, cols in jour_cols.items():
            for i, col in enumerate(cols):
                cell = ws.cell(row, col)
                valeur = cell.value

                if valeur and str(valeur).strip():
                    semaine = i + 1
                    code = str(valeur).strip()
                    code_principal = code.split('/')[0] if '/' in code else code
                    couleur = CODES_COULEURS.get(code_principal, 'FFFFFFFF')

                    repartition_audiences[cle][jour][semaine] = {
                        'code': code,
                        'couleur': couleur
                    }

    wb.close()
    return repartition_audiences